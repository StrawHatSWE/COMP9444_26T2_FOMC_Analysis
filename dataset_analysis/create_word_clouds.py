from __future__ import annotations

import argparse
from pathlib import Path

import matplotlib  
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from wordcloud import WordCloud



def extract_second_column_texts(file_path: Path, data_filter_type: str) -> list[str]:
    """Read sentence text and label values from the first worksheet by scanning headers."""
    workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
    worksheet = workbook.worksheets[0]

    labels = {
        "dovish": 0,
        "hawkish": 1,
        "neutral": 2,
    }

    valid_labels = {0, 1, 2}
    filter_label = None
    if data_filter_type:
        filter_label = labels.get(data_filter_type)
        if filter_label is None:
            try:
                filter_label = int(data_filter_type)
            except (TypeError, ValueError):
                raise ValueError(f"Unknown data_filter_type: {data_filter_type}")

    header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if header_row is None:
        workbook.close()
        return []

    header_map = {str(value).strip().lower(): idx for idx, value in enumerate(header_row) if value is not None}
    label_idx = header_map.get("label")
    text_idx = header_map.get("sentence", 1)

    if label_idx is None:
        workbook.close()
        raise ValueError(f"No 'label' column found in {file_path}")

    texts: list[str] = []
    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if not row:
            continue

        if label_idx >= len(row) or text_idx >= len(row):
            continue

        second_column_value = row[text_idx]
        fourth_column_value = row[label_idx]

        if second_column_value is None or fourth_column_value is None:
            continue

        try:
            label_value = int(fourth_column_value)
        except (TypeError, ValueError):
            continue

        if label_value not in valid_labels:
            continue

        if filter_label is not None and label_value != filter_label:
            continue

        text = str(second_column_value).strip()
        if text:
            texts.append(text)

    workbook.close()
    return texts


def build_word_cloud(input_dir: Path, output_path: Path, data_filter_type: str) -> None:
    xlsx_files = sorted(f for f in input_dir.glob("*.xlsx") if not f.name.startswith("~$"))
    if not xlsx_files:
        raise FileNotFoundError(f"No .xlsx files found in {input_dir}")

    all_texts: list[str] = []
    for xlsx_file in xlsx_files:
        all_texts.extend(extract_second_column_texts(xlsx_file, data_filter_type))

    corpus = " ".join(all_texts)
    if not corpus.strip():
        raise ValueError("No text was found in the second column of the provided Excel files")

    wordcloud = WordCloud(
        width=1400,
        height=900,
        background_color="white",
        max_words=200,
        collocations=False,
    ).generate(corpus)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    plt.figure(figsize=(14, 8))
    plt.imshow(wordcloud, interpolation="bilinear")
    plt.axis("off")
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close()

    print(f"Saved word cloud to {output_path}")
    print(f"Processed {len(xlsx_files)} Excel file(s) and {len(all_texts)} text entry(ies)")


def main() -> None:
    input_dir = Path("C:\\git\\fomc-hawkish-dovish\\training_data\\test-and-training\\training_data").expanduser().resolve()
    output_path = Path(".\\database_analysis\\diagrams\\wordcloud.png").expanduser().resolve()
    data_filter_type = ""

    build_word_cloud(input_dir, output_path, data_filter_type)

    output_path = Path(".\\database_analysis\\diagrams\\wordcloud_hawkish.png").expanduser().resolve()
    data_filter_type = "hawkish"

    build_word_cloud(input_dir, output_path, data_filter_type)

    output_path = Path(".\\database_analysis\\diagrams\\wordcloud_dovish.png").expanduser().resolve()
    data_filter_type = "dovish"

    build_word_cloud(input_dir, output_path, data_filter_type)

    output_path = Path(".\\database_analysis\\diagrams\\wordcloud_neutral.png").expanduser().resolve()
    data_filter_type = "neutral"

    build_word_cloud(input_dir, output_path, data_filter_type)


if __name__ == "__main__":
    main()

