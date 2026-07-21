from __future__ import annotations

import argparse
from pathlib import Path

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from openpyxl import load_workbook


def count_label_occurrences(input_dir: Path) -> dict[str, int]:
    """Count how many times each valid label appears across all Excel files."""
    xlsx_files = sorted(f for f in input_dir.glob("*.xlsx") if not f.name.startswith("~$"))
    if not xlsx_files:
        raise FileNotFoundError(f"No .xlsx files found in {input_dir}")

    label_names = {0: "dovish", 1: "hawkish", 2: "neutral"}
    counts = {name: 0 for name in label_names.values()}

    for xlsx_file in xlsx_files:
        workbook = load_workbook(filename=xlsx_file, read_only=True, data_only=True)
        worksheet = workbook.worksheets[0]

        header_row = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if header_row is None:
            workbook.close()
            continue

        header_map = {str(value).strip().lower(): idx for idx, value in enumerate(header_row) if value is not None}
        label_idx = header_map.get("label")
        if label_idx is None:
            workbook.close()
            continue

        for row in worksheet.iter_rows(min_row=2, values_only=True):
            if not row or label_idx >= len(row):
                continue

            value = row[label_idx]
            if value is None:
                continue

            try:
                label_value = int(value)
            except (TypeError, ValueError):
                continue

            if label_value in label_names:
                counts[label_names[label_value]] += 1

        workbook.close()

    return counts


def build_label_pie_chart(input_dir: Path, output_path: Path, title: str = "Label frequency") -> None:
    counts = count_label_occurrences(input_dir)
    labels = ["dovish", "hawkish", "neutral"]
    values = [counts[label] for label in labels]

    print(f"Label counts: {counts}")

    output_path.parent.mkdir(parents=True, exist_ok=True)
    fig, ax = plt.subplots(figsize=(8, 8))
    ax.pie(
        values,
        labels=labels,
        autopct="%1.1f%%",
        startangle=140,
        colors=["#4C78A8", "#F58518", "#54A24B"],
    )
    ax.set_title(title)
    ax.axis("equal")
    plt.tight_layout()
    plt.savefig(output_path, dpi=300, bbox_inches="tight")
    plt.close(fig)

    print(f"Saved pie chart to {output_path}")
    for label in labels:
        print(f"{label}: {counts[label]}")


REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_INPUT_DIR = REPO_ROOT / "dataset" / "test-and-training" / "augmented_data" / "augmented_train_data"
DEFAULT_OUTPUT_PATH = Path(__file__).resolve().parent / "diagrams" / "label_distribution_augmented_train.png"


def main() -> None:
    parser = argparse.ArgumentParser(description="Pie chart of label distribution across xlsx files in a directory.")
    parser.add_argument("--input-dir", type=Path, default=DEFAULT_INPUT_DIR)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT_PATH)
    parser.add_argument("--title", default="Label frequency (augmented training data)")
    args = parser.parse_args()

    build_label_pie_chart(args.input_dir.resolve(), args.output.resolve(), args.title)


if __name__ == "__main__":
    main()

